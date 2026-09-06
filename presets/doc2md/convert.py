#!/usr/bin/env python3
"""Convert one office/PDF document to Markdown and print a JSON result.

Invoked as a child process by src/doc2md.cjs. Everything it needs to say
travels in the JSON on stdout, so the Node side never has to interpret a
traceback:

    {"ok": true,  "markdown": "...", "note": null, "truncated": false,
     "rows": 0, "pages": 0, "markup_bytes": 0}
    {"ok": false, "reason": "no-text", "detail": "..."}

Exit status is 0 whenever the JSON was written, including for a refusal. A
non-zero exit means the interpreter itself failed and the caller falls back to
letting the original file be read as it always was.

Two things here are defenses rather than features:

* Zip bombs. pptx/xlsx/docx are zip containers, and a hostile attachment can
  declare a small size and expand to fill the disk. The central directory is
  checked first because it is cheap, and then every member is decompressed in
  chunks against a hard ceiling, because the central directory is written by
  whoever built the file and can simply lie.
* Row count. Conversion time tracks spreadsheet rows, not bytes: a 6MB PDF
  converts in about a second while a 6MB 200,000-row workbook takes about a
  minute. Past the row cap the sheet is converted head-first by hand and the
  truncation is stated in the result, because a silently shortened table is
  worse than no table.
"""

import json
import os
import sys
import zipfile

MAX_UNCOMPRESSED = 500 * 1024 * 1024
MAX_RATIO = 200
ROW_CAP = int(os.environ.get("CTS_DOC2MD_ROW_CAP", "50000"))
ZIP_EXTS = {".pptx", ".xlsx", ".docx"}


def fail(reason, detail=""):
    json.dump({"ok": False, "reason": reason, "detail": str(detail)[:500]}, sys.stdout)
    sys.exit(0)


# A password-protected OOXML file is not a zip at all: Office wraps the whole
# package in an OLE compound file whose streams hold the ciphertext. Opening it
# as a zip therefore reports "not a zip file", which reads as a broken download
# and sends the user looking for the wrong problem.
OLE_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def encryption_problem(path, ext):
    """('encrypted', detail) when the file is password-protected, else None."""
    try:
        with open(path, "rb") as fh:
            head = fh.read(8)
    except OSError:
        return None

    # Legacy .xls is an OLE file by design, so the magic alone proves nothing
    # there. For the modern formats it can only mean encryption.
    if ext in ZIP_EXTS and head == OLE_MAGIC:
        return ("encrypted", "password-protected Office file (OLE-wrapped)")

    # Some producers keep the zip container and put the ciphertext inside it.
    if ext in ZIP_EXTS:
        try:
            import zipfile
            with zipfile.ZipFile(path) as z:
                names = z.namelist()
            if any(n.startswith("EncryptedPackage") for n in names):
                return ("encrypted", "password-protected Office file")
        except Exception:
            return None

    if ext == ".pdf":
        try:
            from pdfminer.pdfparser import PDFParser
            from pdfminer.pdfdocument import PDFDocument
            with open(path, "rb") as fh:
                doc = PDFDocument(PDFParser(fh))
            # An empty owner password is the ordinary "printing restricted"
            # case, which extracts fine. Only a document that refuses to open
            # counts as encrypted here.
            if doc.encryption is not None and not doc.is_extractable:
                return ("encrypted", "password-protected PDF")
        except Exception as e:
            if "password" in str(type(e).__name__).lower() or "password" in str(e).lower():
                return ("encrypted", "password-protected PDF")
            return None
    return None


# Enterprise DRM (Fasoo, MarkAny, SoftCamp and the like) does not password a
# document — it wraps the whole file, and only processes the vendor's agent
# has whitelisted ever see plaintext. Python is not one of them, so the bytes
# on disk are ciphertext with a vendor header. Names are matched only to say
# *which* product to go to; the classification does not depend on them.
DRM_MARKERS = [b"FASOO", b"MarkAny", b"MAWDRM", b"SoftCamp", b"Sherpa", b"TrustDRM",
               b"DocuGate", b"WISEDRM", b"UNIDOCS", b"SecureDoc"]
# PDF DRM announces itself as a security handler, and these names are public.
PDF_DRM_FILTERS = [b"FOPN_foweb", b"EBX_HANDLER", b"Adobe.APS", b"FOPN_fLock"]


def drm_hint(head):
    """Vendor name found in the file header, or None."""
    for marker in DRM_MARKERS:
        if marker in head or marker.decode().encode("utf-16-le") in head:
            return marker.decode()
    return None


def wrapper_problem(path, ext):
    """('drm-protected', detail) when the container is not the format at all."""
    try:
        with open(path, "rb") as fh:
            head = fh.read(8192)
    except OSError:
        return None
    if not head:
        return None

    if ext in ZIP_EXTS:
        # A truncated download still starts with a zip local-file header; a
        # wrapped file does not start with anything of the format at all.
        # Telling those two apart is the difference between "re-download it"
        # and "get it released from DRM".
        if head[:2] == b"PK" or head[:8] == OLE_MAGIC:
            return None
        vendor = drm_hint(head)
        return ("drm-protected",
                "DRM-wrapped file (%s)" % vendor if vendor else "the file is not an Office container at all")

    if ext == ".pdf":
        if not head.startswith(b"%PDF"):
            vendor = drm_hint(head)
            return ("drm-protected",
                    "DRM-wrapped file (%s)" % vendor if vendor else "the file is not a PDF at all")
        try:
            with open(path, "rb") as fh:
                blob = fh.read(2_000_000)
        except OSError:
            return None
        for f in PDF_DRM_FILTERS:
            if f in blob:
                return ("drm-protected", "PDF with a DRM security handler (%s)" % f.decode())
    return None


def check_zip(path):
    """Classify an archive before opening it as a document.

    Returns None when it is safe to convert, or a (reason, detail) pair. The
    reason distinguishes a hostile file from a merely broken one: those want
    opposite handling, and calling a truncated download a zip bomb would send
    the user hunting for an attacker who is not there.
    """
    try:
        with zipfile.ZipFile(path) as zf:
            infos = zf.infolist()
            declared = sum(i.file_size for i in infos)
            packed = sum(i.compress_size for i in infos) or 1
            if declared > MAX_UNCOMPRESSED or declared / packed > MAX_RATIO:
                return ("unsafe-archive",
                        "declared size %d bytes at %.0fx compression" % (declared, declared / packed))
            # The numbers above came from the archive itself, so verify them by
            # actually decompressing, stopping the moment the running total
            # passes the ceiling rather than once the disk is full. Understated
            # sizes are caught here twice over: by this budget, and by the
            # CRC-32 check zipfile performs while streaming, which fails as
            # soon as a member's real contents disagree with its header.
            budget = MAX_UNCOMPRESSED
            for info in infos:
                with zf.open(info) as member:
                    while True:
                        chunk = member.read(1 << 20)
                        if not chunk:
                            break
                        budget -= len(chunk)
                        if budget <= 0:
                            return ("unsafe-archive", "expands past %d bytes" % MAX_UNCOMPRESSED)
    except zipfile.BadZipFile as e:
        return ("bad-archive", str(e))
    return None


# Which entries inside a zip container hold the document's own text. The rest
# of the archive is media, themes and relationship tables — bytes a reader
# would never wade through even without a converter.
BODY_XML = {
    ".pptx": ("ppt/slides/", "ppt/notesSlides/"),
    ".docx": ("word/document.xml", "word/footnotes.xml", "word/endnotes.xml"),
    ".xlsx": ("xl/worksheets/", "xl/sharedStrings.xml"),
}


def markup_bytes(path, ext):
    """Uncompressed size of the body markup inside a zip document, or 0.

    This prices the alternative to converting. A model cannot read the binary,
    so the fallback a reader actually reaches for is unzipping the container
    and wading through its XML — where tags and style attributes outweigh the
    text several times over.
    """
    prefixes = BODY_XML.get(ext)
    if not prefixes:
        return 0
    try:
        import zipfile
        with zipfile.ZipFile(path) as z:
            return sum(i.file_size for i in z.infolist()
                       if i.filename.endswith(".xml")
                       and any(i.filename.startswith(p) for p in prefixes))
    except Exception:
        return 0


def pdf_pages(path):
    """Page count of a PDF, or 0 when it cannot be counted.

    The count is what prices the alternative to converting: attaching a PDF
    to a message bills every page as an image, while the conversion bills
    only the extracted text. pdfminer ships with markitdown's pdf extra, so
    this costs no extra dependency.
    """
    try:
        from pdfminer.pdfpage import PDFPage
        with open(path, "rb") as fh:
            return sum(1 for _ in PDFPage.get_pages(fh))
    except Exception:
        return 0


def sheet_rows(path):
    """Total rows across every sheet, or None when openpyxl cannot say."""
    try:
        import openpyxl
    except ImportError:
        return None
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        try:
            return sum(ws.max_row or 0 for ws in wb.worksheets)
        finally:
            wb.close()
    except Exception:
        return None


def md_cell(v):
    if v is None:
        return ""
    return str(v).replace("|", "\\|").replace("\n", " ")


def head_of_workbook(path, cap):
    """Markdown for the first `cap` rows, sheet by sheet.

    Used only past the row cap. markitdown would produce nicer output, but it
    reads the whole workbook first, which is the cost being avoided.
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True)
    out = []
    left = cap
    try:
        for ws in wb.worksheets:
            if left <= 0:
                break
            out.append("## %s" % ws.title)
            header_written = False
            for row in ws.iter_rows(values_only=True):
                if left <= 0:
                    break
                cells = [md_cell(c) for c in row]
                out.append("| " + " | ".join(cells) + " |")
                if not header_written:
                    out.append("| " + " | ".join(["---"] * len(cells)) + " |")
                    header_written = True
                left -= 1
            out.append("")
    finally:
        wb.close()
    return "\n".join(out)


def main():
    if len(sys.argv) < 2:
        fail("usage", "convert.py <file>")
    path = sys.argv[1]
    if not os.path.isfile(path):
        fail("missing", path)

    ext = os.path.splitext(path)[1].lower()

    # Checked before anything else opens the file: an encrypted document is a
    # normal thing to receive, not a failure to report as corruption.
    locked = encryption_problem(path, ext)
    if locked:
        fail(locked[0], locked[1])

    # After the password check, because an OLE-wrapped Office file is an
    # encrypted document rather than a DRM-wrapped one.
    wrapped = wrapper_problem(path, ext)
    if wrapped:
        fail(wrapped[0], wrapped[1])

    if ext in ZIP_EXTS:
        problem = check_zip(path)
        if problem:
            fail(problem[0], problem[1])

    note = None
    truncated = False
    rows = 0
    pages = pdf_pages(path) if ext == ".pdf" else 0
    markup = markup_bytes(path, ext)

    if ext in (".xlsx", ".xls"):
        counted = sheet_rows(path)
        rows = counted or 0
        if counted and counted > ROW_CAP:
            try:
                text = head_of_workbook(path, ROW_CAP)
            except Exception as e:
                fail("convert-failed", e)
            note = ("전체 %d행 가운데 앞 %d행만 변환했습니다. "
                    "전수 분석이 필요하면 원본을 직접 다루십시오." % (counted, ROW_CAP))
            json.dump({"ok": True, "markdown": text, "note": note,
                       "truncated": True, "rows": counted, "pages": 0,
                       "markup_bytes": markup}, sys.stdout)
            return

    try:
        from markitdown import MarkItDown
    except ImportError as e:
        fail("no-markitdown", e)

    try:
        result = MarkItDown().convert(path)
        text = (result.text_content or "").strip()
    except Exception as e:
        # markitdown surfaces the password failure from whichever backend hit
        # it, so the type name is the reliable part.
        blob = (type(e).__name__ + " " + str(e)).lower()
        if "password" in blob or "encrypted" in blob:
            fail("encrypted", "the file is password-protected")
        fail("convert-failed", e)

    if not text:
        # An empty file would read to the model as a document with nothing in
        # it, which is a different and worse claim than "could not extract".
        fail("no-text", "converter returned nothing")

    json.dump({"ok": True, "markdown": text, "note": note,
               "truncated": truncated, "rows": rows, "pages": pages,
               "markup_bytes": markup}, sys.stdout)


if __name__ == "__main__":
    main()
